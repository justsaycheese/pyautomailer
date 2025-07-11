import logging
import os
import random
import smtplib
import re, uuid, os
import mimetypes
import sys
import threading
import time
import json
import pythoncom
from email import encoders
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from tkinter import (
    END,
    Button,
    Entry,
    Frame,
    Label,
    OptionMenu,
    Scrollbar,
    StringVar,
    Text,
    Tk,
    Toplevel,
    filedialog,
    messagebox,
    scrolledtext,
    ttk,
)

import extract_msg
import pandas as pd
import RTFDE.text_extraction as rtf_te
import win32com.client as win32
from openpyxl import load_workbook

# ─────────────────────────────
# ⚙️ Config & Log
# ─────────────────────────────
# 預設結尾詞清單，不再做為全域變數修改
DEFAULT_CLOSING_STATEMENTS = [
    "Thanks & Best Regards",
    "Kind Regards",
    "Sincerely",
    "With sincere appreciation",
    "With gratitude",
    "Gratefully",
    "Warm regards",
]
DELAY_SEND = 10
DELAY_DRAFT = 1
LOG_FILE = "automailer_log.txt"
logging.basicConfig(
    filename=LOG_FILE, level=logging.INFO, format="%(asctime)s - %(message)s"
)


def patch_rtfde_decode() -> None:
    """Patch RTFDE to ignore undecodable hex characters."""

    def _patched_decode_hex_char(item: bytes, codec: str | None):
        if codec is None:
            codec = "CP1252"
        try:
            return item.decode(codec).encode()
        except UnicodeDecodeError:
            return item.decode(codec, errors="ignore").encode()

    rtf_te.decode_hex_char = _patched_decode_hex_char


patch_rtfde_decode()

# Excel 全表識別
ALL_SHEETS = "整本活頁簿"

# ─────────────────────────────
# 📨 Backend Classes
# ─────────────────────────────


class EmailBackend:
    """Email backend base class."""

    def send(
        self,
        mode: str,
        recipient: str,
        subject: str,
        html_body: str,
        embedded_images: dict[str, Path],
        attachments: list[Path],
    ) -> None:
        raise NotImplementedError


class OutlookBackend(EmailBackend):
    def __init__(self, account_name: str | None = None):
        self.outlook = win32.Dispatch("Outlook.Application")
        session = self.outlook.GetNamespace("MAPI")
        self.account = None
        if account_name:
            for acct in session.Accounts:
                if acct.DisplayName == account_name:
                    self.account = acct
                    break

    def send(
        self,
        mode: str,
        recipient: str,
        subject: str,
        html_body: str,
        embedded_images: dict[str, Path],
        attachments: list[Path],
    ) -> None:
        mail = self.outlook.CreateItem(0)
        if self.account:
            mail._oleobj_.Invoke(64209, 0, 8, 1, self.account)

        mail.BodyFormat = 2          # olFormatHTML，先指定格式
        mail.Subject = subject
        mail.To = recipient

        # 1️⃣ 先附上要嵌入的圖片
        for cid, path in embedded_images.items():
            attach = mail.Attachments.Add(
                Source=str(path),
                Type=1,    # olByValue
                Position=0
            )
            pa = attach.PropertyAccessor
            pa.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)   # CID
            pa.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x7FFE000B", True)  # PR_ATTACHMENT_HIDDEN

        # 2️⃣ 再放進 HTML 內容
        mail.HTMLBody = html_body

        # 3️⃣ 其餘一般附件最後附
        for file_path in attachments:
            mail.Attachments.Add(Source=str(file_path))

        if mode == "send":
            mail.Send()
        else:
            mail.Save()



class SmtpBackend(EmailBackend):
    def __init__(self, host: str, port: int, username: str, password: str):
        self.host = host
        self.port = port
        self.username = username
        self.password = password

    def send(
        self,
        mode: str,
        recipient: str,
        subject: str,
        html_body: str,
        embedded_images: dict[str, Path],
        attachments: list[Path],
    ) -> None:
        msg_root = MIMEMultipart("related")
        msg_root["Subject"] = subject
        msg_root["From"] = self.username
        msg_root["To"] = recipient
        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(html_body, "html", "utf-8"))
        msg_root.attach(alt)

        for cid, path in embedded_images.items():
            with open(path, "rb") as f:
                data = f.read()
            mime_type, _ = mimetypes.guess_type(path)
            if mime_type and mime_type.startswith("image/"):
                _, subtype = mime_type.split("/", 1)
            else:
                subtype = path.suffix.lstrip(".") or "png"
            img = MIMEImage(data, _subtype=subtype)
            img.add_header("Content-ID", f"<{cid}>")
            msg_root.attach(img)

        for file_path in attachments:
            with open(file_path, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition", "attachment", filename=file_path.name
            )
            msg_root.attach(part)

        if mode == "draft":
            draft_dir = get_base_dir() / "drafts"
            draft_dir.mkdir(exist_ok=True)
            with open(draft_dir / f"{recipient}.eml", "w", encoding="utf-8") as f:
                f.write(msg_root.as_string())
            return

        with smtplib.SMTP(self.host, self.port) as server:
            server.starttls()
            server.login(self.username, self.password)
            server.sendmail(self.username, [recipient], msg_root.as_string())


# ─────────────────────────────
# 📂 Utils
# ─────────────────────────────

def get_excel_sheets(file_path: str) -> list[str]:
    """取得 Excel 檔案的所有工作表名稱。若非 Excel 則回傳空清單。"""
    if Path(file_path).suffix.lower() in [".xls", ".xlsx"]:
        try:
            wb = load_workbook(file_path, read_only=True)
            return wb.sheetnames
        except Exception:
            return []
    return []


def load_recipients_or_csv(file_path, visible_only=False, sheet_name=None):
    ext = Path(file_path).suffix.lower()
    if ext == ".csv":
        df = pd.read_csv(file_path)
    elif ext in [".xls", ".xlsx"]:
        if not visible_only:
            if sheet_name == ALL_SHEETS:
                xls = pd.ExcelFile(file_path)
                df_list = [pd.read_excel(xls, sh) for sh in xls.sheet_names]
                df = pd.concat(df_list, ignore_index=True)
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            wb = load_workbook(file_path, data_only=True)
            sheets = (
                wb.sheetnames
                if sheet_name == ALL_SHEETS
                else [sheet_name or wb.sheetnames[0]]
            )
            all_rows = []
            headers = None
            for sh in sheets:
                ws = wb[sh]
                if headers is None:
                    headers = [cell.value for cell in ws[1]]
                visible_rows = [
                    [cell.value for cell in row]
                    for row in ws.iter_rows(min_row=2)
                    if not ws.row_dimensions[row[0].row].hidden
                ]
                all_rows.extend(visible_rows)
            df = pd.DataFrame(all_rows, columns=headers)
    else:
        raise ValueError(f"Unsupported file type: {file_path}")

    return df


def validate_recipient_columns(df):
    """Ensure required columns are present in the loaded DataFrame."""
    required = {"Email", "Salutation"}
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Missing column(s): {', '.join(missing)}")


def get_base_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    else:
        return Path(__file__).parent

SETTINGS_FILE = get_base_dir() / "settings.json"

def load_settings_file():
    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Failed to load settings: {e}")
    return {}

def save_settings_file(data):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logging.error(f"Failed to save settings: {e}")
    
def safe_cid(stem: str) -> str:
    """
    將檔名主體轉成僅含大小寫英數、底線、連字號的安全 CID。
    再加 8 碼隨機碼，確保全域唯一。
    """
    clean = re.sub(r'[^A-Za-z0-9_-]+', '_', stem)      # 空白、()、中文 → _
    return f"{uuid.uuid4().hex[:8]}_{clean}"

def load_embeds(embed_dir):
    """載入要嵌入的圖片，回傳 {cid: Path} 的 dict。"""
    if embed_dir is None:
        return {}
    embed_dir = Path(embed_dir)
    return {
        safe_cid(f.stem): f.resolve()
        for f in embed_dir.glob("*")
        if f.suffix.lower() in [".png", ".jpg", ".jpeg", ".gif"]
    }


def load_attachments(attachment_dir):
    """載入附件檔案，回傳 Path list。"""
    if attachment_dir is None:
        return []
    attachment_dir = Path(attachment_dir)
    return [f.resolve() for f in attachment_dir.glob("*") if f.is_file()]


def generate_image_html(embeds):
    return "".join(
        f'<img src="cid:{cid}" style="display:block; margin-bottom:10px;"><br>'
        for cid in embeds
    )


# ─────────────────────────────
# 🖥️ GUI Class
# ─────────────────────────────
class GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Automailer 自動寄信工具")

        # 模式切換（寄送或存稿）
        self.mode_var = StringVar(value="draft")
        self.recipient_file = ""
        self.exclusion_file = ""
        self.msg_template = ""
        self.embed_dir = None
        self.attachment_dir = None

        self.embed_files = StringVar(value="尚未選擇")
        self.attachment_files = StringVar(value="尚未選擇")
        self.recipient_label = StringVar(value="尚未選擇")
        self.exclusion_label = StringVar(value="尚未選擇")
        self.template_label = StringVar(value="尚未選擇")
        self.select_mode_var = StringVar(value="資料夾")  # 預設「資料夾」模式
        self.folder_mode = True  # 與前保持一致的布林旗標
        self.embed_paths = {}  # 字典
        self.attachments = []
        self.recipient_sheet_var = StringVar(value=ALL_SHEETS)
        self.exclusion_sheet_var = StringVar(value=ALL_SHEETS)

        # 進度文字和 Progressbar
        self.progress_label = StringVar(value="")
        self.progress_bar = None

        # 日誌視窗相關
        self.log_window = None
        self.log_buffer = ["✅ 程式已啟動"]

        # ─── pause_event & cancel_event ───
        self.pause_event = threading.Event()
        self.pause_event.set()  # 一開始為「已設定」，代表不暫停

        self.cancel_event = threading.Event()  # 一開始為 False，代表未取消

        # ——取得 Outlook Accounts ——
        outlook_app = win32.Dispatch("Outlook.Application")
        session = outlook_app.GetNamespace("MAPI")
        accounts = [acct.DisplayName for acct in session.Accounts]

        # 如果只有一個帳戶，也把它放進去
        if not accounts:
            accounts = ["(No Account Found)"]
        self.account_var = StringVar(root)
        self.backend_var = StringVar(value="Outlook")
        self.smtp_host = StringVar(value="")
        self.smtp_port = StringVar(value="587")
        self.smtp_user = StringVar(value="")
        self.smtp_pass = StringVar(value="")

        # 讀取設定檔並套用
        cfg = load_settings_file()
        self.mode_var.set(cfg.get("mode", self.mode_var.get()))
        self.backend_var.set(cfg.get("backend", self.backend_var.get()))
        self.select_mode_var.set(cfg.get("select_mode", self.select_mode_var.get()))
        self.folder_mode = self.select_mode_var.get() == "資料夾"
        acc = cfg.get("account")
        self.account_var.set(acc if acc in accounts else accounts[0])
        self.smtp_host.set(cfg.get("smtp_host", ""))
        self.smtp_port.set(cfg.get("smtp_port", "587"))
        self.smtp_user.set(cfg.get("smtp_user", ""))
        self.smtp_pass.set(cfg.get("smtp_pass", ""))
        self.recipient_file = cfg.get("recipient_file", "")
        if self.recipient_file:
            self.recipient_label.set(Path(self.recipient_file).name)
        self.recipient_sheet_var.set(cfg.get("recipient_sheet", ALL_SHEETS))
        self.exclusion_file = cfg.get("exclusion_file", "")
        if self.exclusion_file:
            self.exclusion_label.set(Path(self.exclusion_file).name)
        self.exclusion_sheet_var.set(cfg.get("exclusion_sheet", ALL_SHEETS))
        self.msg_template = cfg.get("msg_template", "")
        if self.msg_template:
            self.template_label.set(Path(self.msg_template).name)
        if self.folder_mode:
            embed_dir = cfg.get("embed_dir")
            if embed_dir:
                self.embed_dir = Path(embed_dir)
                self.embed_paths = load_embeds(self.embed_dir)
                self.embed_files.set(", ".join(p.name for p in self.embed_paths.values()) or "無檔案")
            attachment_dir = cfg.get("attachment_dir")
            if attachment_dir:
                self.attachment_dir = Path(attachment_dir)
                self.attachments = load_attachments(self.attachment_dir)
                self.attachment_files.set(", ".join(p.name for p in self.attachments) or "無檔案")
        else:
            embed_list = cfg.get("embed_files", [])
            if embed_list:
                self.embed_paths = {safe_cid(Path(p).stem): Path(p) for p in embed_list}
                self.embed_files.set(", ".join(Path(p).name for p in embed_list))
            attach_list = cfg.get("attachment_files", [])
            if attach_list:
                self.attachments = [Path(p) for p in attach_list]
                self.attachment_files.set(", ".join(Path(p).name for p in self.attachments))
        saved_closing = cfg.get("closing_statements")

        # ──────────────── UI Frames ────────────────
        mode_frame = Frame(root, pady=5, padx=5, relief="groove", borderwidth=2)
        mode_frame.grid(row=0, column=0, columnspan=2, sticky="EW")
        self.account_label = Label(mode_frame, text="寄件帳戶：")
        self.account_label.grid(row=0, column=0, sticky="w", pady=5)

        self.account_menu = OptionMenu(mode_frame, self.account_var, *accounts)
        self.account_menu.grid(row=0, column=1, sticky="W", pady=5)
        self.account_menu.config(width=20)

        Label(mode_frame, text="寄信後端:").grid(row=1, column=0, sticky="W")
        backend_menu = OptionMenu(
            mode_frame,
            self.backend_var,
            "Outlook",
            "SMTP",
            command=self.on_backend_change,
        )
        backend_menu.config(width=7)
        backend_menu.grid(row=1, column=1, sticky="W")

        Label(mode_frame, text="選擇寄送模式:").grid(row=2, column=0, sticky="W")
        mode_menu = OptionMenu(mode_frame, self.mode_var, "send", "draft")
        mode_menu.config(width=5)
        mode_menu.grid(row=2, column=1, sticky="W")

        self.smtp_frame = Frame(root, pady=5, padx=5, relief="groove", borderwidth=2)
        self.smtp_frame.grid(row=1, column=0, columnspan=2, sticky="EW")
        Label(self.smtp_frame, text="SMTP 主機:").grid(row=0, column=0, sticky="W")
        Entry(self.smtp_frame, textvariable=self.smtp_host, width=25).grid(
            row=0, column=1, sticky="W"
        )
        Label(self.smtp_frame, text="Port:").grid(row=0, column=2, sticky="W")
        Entry(self.smtp_frame, textvariable=self.smtp_port, width=5).grid(
            row=0, column=3, sticky="W"
        )
        Label(self.smtp_frame, text="User:").grid(row=1, column=0, sticky="W")
        Entry(self.smtp_frame, textvariable=self.smtp_user, width=25).grid(
            row=1, column=1, sticky="W"
        )
        Label(self.smtp_frame, text="Password:").grid(row=1, column=2, sticky="W")
        Entry(self.smtp_frame, textvariable=self.smtp_pass, show="*", width=10).grid(
            row=1, column=3, sticky="W"
        )
        self.smtp_frame.grid_remove()

        file_frame = Frame(root, pady=5, padx=5, relief="groove", borderwidth=2)
        file_frame.grid(row=2, column=0, columnspan=2, sticky="EW")
        self.embed_btn = Button(
            file_frame, text="🖼 選擇圖片資料夾", command=self.select_embed, width=20
        )
        self.embed_btn.grid(row=1, column=0, pady=5)
        Label(
            file_frame, textvariable=self.embed_files, wraplength=270, justify="left"
        ).grid(row=1, column=1, sticky="W")

        self.attachment_btn = Button(
            file_frame,
            text="📎 選擇附件資料夾",
            command=self.select_attachment,
            width=20,
        )
        self.attachment_btn.grid(row=2, column=0, pady=5)
        Label(
            file_frame,
            textvariable=self.attachment_files,
            wraplength=270,
            justify="left",
        ).grid(row=2, column=1, sticky="W")

        inner_frame1 = Frame(file_frame, borderwidth=0)
        inner_frame1.grid(row=0, column=0, columnspan=2, sticky="EW")
        Label(inner_frame1, text="選取模式：").grid(row=0, column=0, sticky="W", pady=5)
        mode_select = OptionMenu(
            inner_frame1,
            self.select_mode_var,
            "資料夾",
            "多檔案",
            command=self.on_select_mode,
        )
        mode_select.config(width=8)
        mode_select.grid(row=0, column=1, sticky="W", pady=5)

        choose_frame = Frame(root, pady=5, padx=5, relief="groove", borderwidth=2)
        choose_frame.grid(row=3, column=0, columnspan=2, sticky="EW")
        Button(
            choose_frame, text="📋 選擇收件人", command=self.load_recipients, width=20
        ).grid(row=0, column=0, pady=5)
        Label(
            choose_frame,
            textvariable=self.recipient_label,
            wraplength=270,
            justify="left",
        ).grid(row=0, column=1, sticky="W")
        self.recipient_sheet_menu = OptionMenu(
            choose_frame, self.recipient_sheet_var, self.recipient_sheet_var.get()
        )
        self.recipient_sheet_menu.config(width=8)
        self.recipient_sheet_menu.grid(row=0, column=2, sticky="W")
        Button(
            choose_frame, text="🚫 選擇排除清單", command=self.load_exclusions, width=20
        ).grid(row=1, column=0, pady=5)
        Label(
            choose_frame,
            textvariable=self.exclusion_label,
            wraplength=270,
            justify="left",
        ).grid(row=1, column=1, sticky="W")
        self.exclusion_sheet_menu = OptionMenu(
            choose_frame, self.exclusion_sheet_var, self.exclusion_sheet_var.get()
        )
        self.exclusion_sheet_menu.config(width=8)
        self.exclusion_sheet_menu.grid(row=1, column=2, sticky="W")

        if self.recipient_file:
            self.update_sheet_menu(
                self.recipient_sheet_menu,
                self.recipient_sheet_var,
                get_excel_sheets(self.recipient_file),
            )
        if self.exclusion_file:
            self.update_sheet_menu(
                self.exclusion_sheet_menu,
                self.exclusion_sheet_var,
                get_excel_sheets(self.exclusion_file),
            )
        Button(
            choose_frame,
            text="✉ 選擇郵件範本",
            command=self.load_msg_template,
            width=20,
        ).grid(row=2, column=0, pady=5)
        Label(
            choose_frame,
            textvariable=self.template_label,
            wraplength=270,
            justify="left",
        ).grid(row=2, column=1, sticky="W")

        Button(root, text="🚀 開始寄信", command=self.start_process).grid(
            row=4, column=0, pady=10
        )
        Button(root, text="🪵 查看日誌", command=self.show_log_window).grid(
            row=4, column=1
        )

        Label(root, text="結尾詞 (一行一個)").grid(row=5, column=0, columnspan=2)
        self.closing_text = scrolledtext.ScrolledText(root, height=7, width=50)
        self.closing_text.grid(row=6, column=0, columnspan=2)
        if saved_closing:
            self.closing_text.insert(END, "\n".join(saved_closing))
            self.closing_statements = saved_closing
        else:
            self.closing_text.insert(END, "\n".join(DEFAULT_CLOSING_STATEMENTS))
            self.closing_statements = DEFAULT_CLOSING_STATEMENTS

        # ─── Pause/Resume 按鈕 & Cancel 按鈕（一開始先放位置，再隱藏） ───
        self.pause_button = Button(
            root,
            text="暫停",
            font=("Arial", 12, "bold"),
            fg="#f00",
            state="normal",
            command=self.toggle_pause,
            width=10,
            height=1,
        )
        self.pause_button.grid(row=8, column=0, pady=5)
        self.pause_button.grid_remove()  # 先隱藏

        self.cancel_button = Button(
            root,
            text="取消",
            font=("Arial", 12, "bold"),
            fg="#f00",
            state="normal",
            command=self.cancel_process,
            width=10,
            height=1,
        )
        self.cancel_button.grid(row=8, column=1, pady=5)
        self.cancel_button.grid_remove()  # 先隱藏

        # ────────────── 進度區塊 ──────────────
        Label(root, textvariable=self.progress_label).grid(
            row=10, column=0, columnspan=2, pady=5, sticky="S"
        )
        self.progress_bar = ttk.Progressbar(root, length=300, mode="determinate")
        self.progress_bar.grid(row=11, column=0, columnspan=2, pady=5)

        self.save_button = Button(root, text="💾 儲存設定", command=self.save_settings)
        self.save_button.grid(row=12, column=0, columnspan=2, pady=5)

    def on_select_mode(self, choice):
        """當 OptionMenu 變動時呼叫；同步更新 folder_mode 與按鈕文字"""
        self.folder_mode = choice == "資料夾"  # True=資料夾模式

        if self.folder_mode:
            # ── 切到「資料夾」──
            # 清掉多檔案模式用到的屬性
            self.embed_paths.clear()
            self.attachments.clear()
            # 也把顯示文字歸零
            self.embed_files.set("尚未選擇")
            self.attachment_files.set("尚未選擇")
            # 同時把舊的資料夾路徑重置，讓使用者重新選
            self.embed_dir = None
            self.attachment_dir = None

            # 更新按鈕文字
            self.embed_btn.config(text="🖼 選擇圖片資料夾")
            self.attachment_btn.config(text="📎 選擇附件資料夾")
        else:
            # ── 切到「多檔案」──
            # 清掉資料夾模式遺留的屬性
            self.embed_dir = None
            self.attachment_dir = None
            # 同步把顯示文字歸零
            self.embed_files.set("尚未選擇")
            self.attachment_files.set("尚未選擇")
            # 也把多檔案內容清空（防止舊清單殘留）
            self.embed_paths.clear()
            self.attachments.clear()

            # 更新按鈕文字
            self.embed_btn.config(text="🖼 選擇圖片檔案")
            self.attachment_btn.config(text="📎 選擇附件檔案")
            self.log(f"🔀 已切換到 «{choice}» 模式")

    def on_backend_change(self, choice):
        """切換寄信後端時顯示或隱藏 SMTP 設定欄位"""
        if choice == "SMTP":
            self.smtp_frame.grid()
            self.account_menu.grid_remove()
            self.account_label.grid_remove()
        else:
            self.smtp_frame.grid_remove()
            self.account_menu.grid()
            self.account_label.grid()

    def select_embed(self):
        if self.folder_mode:  # ▸ 資料夾模式
            path = filedialog.askdirectory(title="選擇嵌入圖片資料夾")
            if not path:
                return
            self.embed_dir = Path(path)
            self.embed_paths = load_embeds(self.embed_dir)
        else:  # ▸ 多檔案模式
            paths = filedialog.askopenfilenames(
                title="選擇圖片檔案",
                filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif")],
            )
            if not paths:
                return
            # 產生跟 load_embeds 相同結構的 dict
            self.embed_paths = {
                safe_cid(Path(p).stem): Path(p)                  
                for p in paths
            }

        file_names = [p.name for p in self.embed_paths.values()]
        self.embed_files.set(", ".join(file_names) or "無檔案")
        self.log(f"✅ 已載入 {len(self.embed_paths)} 張圖片")

    def select_attachment(self):
        if self.folder_mode:
            # ▸ 資料夾模式
            path = filedialog.askdirectory(title="選擇附件資料夾")
            if not path:
                return
            self.attachment_dir = Path(path)
            self.attachments = load_attachments(self.attachment_dir)
        else:
            # ▸ 多檔案模式
            paths = filedialog.askopenfilenames(title="選擇附件檔案")
            if not paths:
                return
            self.attachments = [Path(p) for p in paths]

        file_names = [p.name for p in self.attachments]
        self.attachment_files.set(", ".join(file_names) or "無檔案")
        self.log(f"✅ 已載入 {len(self.attachments)} 個附件")

    def update_sheet_menu(self, menu, var, sheets):
        menu["menu"].delete(0, "end")
        options = [ALL_SHEETS] + sheets if sheets else [ALL_SHEETS]
        for s in options:
            menu["menu"].add_command(label=s, command=lambda v=s: var.set(v))
        var.set(ALL_SHEETS)

    def load_recipients(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv")]
        )
        if path:
            self.recipient_file = path
            self.recipient_label.set(Path(path).name)
            sheets = get_excel_sheets(path)
            self.update_sheet_menu(self.recipient_sheet_menu, self.recipient_sheet_var, sheets)

    def load_exclusions(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv")]
        )
        if path:
            self.exclusion_file = path
            self.exclusion_label.set(Path(path).name)
            sheets = get_excel_sheets(path)
            self.update_sheet_menu(self.exclusion_sheet_menu, self.exclusion_sheet_var, sheets)

    def load_msg_template(self):
        path = filedialog.askopenfilename(filetypes=[("MSG Files", "*.msg")])
        if path:
            self.msg_template = path
            self.template_label.set(Path(path).name)

    def show_log_window(self):
        if self.log_window and self.log_window.winfo_exists():
            self.log_window.lift()
            return
        self.log_window = Toplevel(self.root)
        self.log_window.title("🪵 日誌紀錄")

        frame = Frame(self.log_window)
        frame.pack(fill="both", expand=True)

        self.log_text = Text(frame, wrap="word", font=("Courier", 9))
        scrollbar = Scrollbar(frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        for msg in self.log_buffer:
            self.log_text.insert(END, msg + "\n")
        self.log_text.config(state="disabled")

        clear_btn = Button(
            frame,
            text="🧹清空",
            command=self.clear_log,
            relief="groove",
            bg="lightgray",
        )
        clear_btn.place(relx=1.0, rely=0.0, anchor="ne", x=1, y=0)

    def clear_log(self):
        self.log_buffer.clear()
        if (
            self.log_window
            and hasattr(self, "log_text")
            and self.log_window.winfo_exists()
        ):
            self.log_text.config(state="normal")
            self.log_text.delete("1.0", END)
            self.log_text.config(state="disabled")
        with open(LOG_FILE, "w") as f:
            f.write("")
        self.log("🧹 日誌已清空")

    def log(self, msg):
        logging.info(msg)
        self.log_buffer.append(msg)
        if (
            self.log_window
            and hasattr(self, "log_text")
            and self.log_window.winfo_exists()
        ):

            def append_log():
                self.log_text.config(state="normal")
                self.log_text.insert(END, msg + "\n")
                self.log_text.see(END)
                self.log_text.config(state="disabled")

            self.root.after(0, append_log)

    def start_process(self):
        # ─── 重新開始時，要先重置進度標籤與進度條 ───
        self.progress_label.set("")
        self.progress_bar["value"] = 0

        user_input = self.closing_text.get("1.0", END).strip().splitlines()
        self.closing_statements = [line.strip() for line in user_input if line.strip()]
        if not self.recipient_file or not self.msg_template:
            messagebox.showerror("錯誤", "請選擇收件人清單和郵件範本")
            return

        # -------- 依模式整理圖片與附件清單 --------
        if self.embed_paths:  # ↖ 多檔案模式
            embedded_images = self.embed_paths
        elif self.embed_dir is not None:  # ↖ 資料夾模式
            embedded_images = load_embeds(self.embed_dir)
        else:
            embedded_images = {}

        if self.attachments:  # ↖ 多檔案模式
            real_attachments = self.attachments
        elif self.attachment_dir is not None:  # ↖ 資料夾模式
            real_attachments = load_attachments(self.attachment_dir)
        else:
            real_attachments = []

        embed_list = (
            "\n".join([f"- {cid} → {p.name}" for cid, p in embedded_images.items()])
            or "無"
        )
        attachment_list = "\n".join([f"- {p.name}" for p in real_attachments]) or "無"
        statement_list = "\n".join(self.closing_statements)
        if self.backend_var.get() == "Outlook":
            account_disp = self.account_var.get()
        else:
            account_disp = self.smtp_user.get()
        confirm_message = f"""📂 檢查完畢，準備寄信：{self.mode_var.get()}

寄件帳戶：{account_disp}
寄件後端：{self.backend_var.get()}

嵌入圖片:
{embed_list}

附加檔案:
{attachment_list}

結尾詞:
{statement_list}

是否確認開始寄信？
"""
        if not messagebox.askyesno("確認寄信", confirm_message):
            self.progress_label.set("⛔ 操作取消")
            self.log("⛔ 操作取消")
            return

        # 在開始之前，重置 cancel_event 並設定 pause_event
        self.cancel_event.clear()
        self.pause_event.set()

        self.log(
            f"📧 寄件帳戶：{account_disp} / 寄件後端：{self.backend_var.get()}"
        )

        # 顯示暫停和取消按鈕
        self.pause_button.grid()  # 從隱藏狀態恢復
        self.pause_button.config(text="暫停")
        self.cancel_button.grid()  # 從隱藏狀態恢復
        self.save_button.grid_remove()

        # 啟動背景執行緒，傳入 pause_event 和 cancel_event
        threading.Thread(
            target=run_automailer,
            args=(
                self.mode_var.get(),
                self.recipient_file,
                self.exclusion_file,
                self.recipient_sheet_var.get(),
                self.exclusion_sheet_var.get(),
                self.msg_template,
                self.update_progress,
                self.log,
                embedded_images,  # ← 改傳「最終 dict」
                real_attachments,  # ← 改傳「最終 list」
                self.pause_event,
                self.cancel_event,
                self.on_finish,
                self.account_var.get(),
                self.backend_var.get(),
                self.smtp_host.get(),
                self.smtp_port.get(),
                self.smtp_user.get(),
                self.smtp_pass.get(),
                self.closing_statements,
            ),
            daemon=True,
        ).start()

    def toggle_pause(self):
        """切換暫停 / 繼續 狀態，並更新按鈕文字。"""
        if self.pause_event.is_set():
            # 由「可執行」變成「暫停」
            self.pause_event.clear()
            self.pause_button.config(text="繼續")
            self.log("⏸️ 已暫停寄送")
        else:
            # 由「暫停」變成「可執行」
            self.pause_event.set()
            self.pause_button.config(text="暫停")
            self.log("▶️ 已繼續寄送")

    def cancel_process(self):
        """使用者點「取消」時，觸發 cancel_event，並隱藏按鈕。"""
        # 設定 cancel_event，讓 run_automailer 迴圈跳出
        self.cancel_event.set()
        self.log("❌ 使用者已取消寄送")
        # 隱藏按鈕
        self.pause_button.grid_remove()
        self.cancel_button.grid_remove()
        # 更新進度文字為已取消
        self.progress_label.set("❌ 已取消寄送")

    def save_settings(self):
        data = {
            "mode": self.mode_var.get(),
            "backend": self.backend_var.get(),
            "account": self.account_var.get(),
            "select_mode": self.select_mode_var.get(),
            "smtp_host": self.smtp_host.get(),
            "smtp_port": self.smtp_port.get(),
            "smtp_user": self.smtp_user.get(),
            "smtp_pass": self.smtp_pass.get(),
            "recipient_file": self.recipient_file,
            "exclusion_file": self.exclusion_file,
            "recipient_sheet": self.recipient_sheet_var.get(),
            "exclusion_sheet": self.exclusion_sheet_var.get(),
            "msg_template": self.msg_template,
            "closing_statements": self.closing_text.get("1.0", END).strip().splitlines(),
        }
                # 根據目前的「選取模式」決定要寫哪一組鍵
        if self.folder_mode:
            # ㈠ 資料夾模式 → 只保留 _dir，完全省略 *_files
            data["embed_dir"] = str(self.embed_dir or "")
            data["attachment_dir"] = str(self.attachment_dir or "")
        else:
            # ㈡ 多檔案模式 → 只保留 *_files，完全省略 _dir
            data["embed_files"] = [str(p) for p in self.embed_paths.values()]
            data["attachment_files"] = [str(p) for p in self.attachments]
            
        save_settings_file(data)
        self.log("✅ 設定已儲存")
        messagebox.showinfo("設定", "設定已儲存")

    def update_progress(self, index, total, current_email):
        pct = int((index + 1) / total * 100)
        self.progress_label.set(f"{pct}% - 處理 {index + 1}/{total}: {current_email}")
        self.progress_bar["value"] = pct
        self.root.update_idletasks()

    def on_finish(self, last_index, total):
        """流程跑完後，把暫停與取消按鈕隱藏掉。"""
        self.pause_button.grid_remove()
        self.cancel_button.grid_remove()
        self.save_button.grid()
        # 可以更新進度文字表達「已完成」：
        finished_count = last_index + 1 if last_index is not None else total
        self.progress_label.set(f"✅ 全部寄送完成 {finished_count}/{total}")
        
# ─────────────────────────────
# 🚀 Email Sending Logic
# ─────────────────────────────
def run_automailer(
    mode,
    recipients_path,
    exclusion_path,
    recipients_sheet,
    exclusion_sheet,
    msg_template_path,
    progress_update,
    logger,
    embedded_images,
    real_attachments,
    pause_event,
    cancel_event,
    finish_callback,
    send_account_name,
    backend_type,
    smtp_host,
    smtp_port,
    smtp_user,
    smtp_pass,
    closing_statements,
):

    use_outlook = backend_type != "SMTP"
    if use_outlook:
        pythoncom.CoInitialize()
        backend = OutlookBackend(send_account_name)
    else:
        backend = SmtpBackend(smtp_host, int(smtp_port or 0), smtp_user, smtp_pass)

    try:
        recipients = load_recipients_or_csv(
            recipients_path,
            visible_only=True,
            sheet_name=recipients_sheet if recipients_sheet != ALL_SHEETS else ALL_SHEETS,
        )
        validate_recipient_columns(recipients)
    except ValueError as e:
        messagebox.showerror("檔案錯誤", str(e))
        logger(f"收件人清單錯誤: {e}")
        if finish_callback:
            finish_callback(None, 0)
        return
    exclusion_emails = []
    if exclusion_path and os.path.exists(exclusion_path):
        try:
            exclusion_df = load_recipients_or_csv(
                exclusion_path, sheet_name=exclusion_sheet if exclusion_sheet != ALL_SHEETS else ALL_SHEETS
            )
            exclusion_emails = exclusion_df["Email"].tolist()
        except Exception as e:
            logger(f"排除清單讀取失敗: {e}")
    filtered = recipients[~recipients["Email"].isin(exclusion_emails)]

    msg = extract_msg.Message(msg_template_path)
    subject = msg.subject
    try:
        raw_html_body = msg.htmlBody
    except UnicodeDecodeError as e:
        logger(f"HTML 解析失敗: {e}")
        raw_html_body = None
    html_body = (
        raw_html_body.decode("utf-8", errors="ignore")
        if isinstance(raw_html_body, bytes)
        else (raw_html_body or "")
    )

    cid_list = list(embedded_images)
    image_html_all = generate_image_html(cid_list)

    total = len(filtered)

    """
    新增參數 cancel_event。每次迴圈開始前或 pause 時，都要檢查 cancel_event 
    是否已被設置。設置就直接結束整個流程。
    新增一個參數 pause_event：threading.Event 物件。
    在每次實際要發送/存稿之前，都先呼叫 pause_event.wait()。
    當 pause_event 被 clear 時，wait() 會阻塞；被 set 時繼續執行。
    """
    last_index = None
    for i, row in filtered.iterrows():
        last_index = i
        # 若使用者按了「取消」，就直接跳出
        if cancel_event.is_set():
            logger("❌ 停止寄送，使用者已取消")
            break

        # 暫停處理：若 pause_event 沒被 set，就持續小睡並每次檢查 cancel_event
        while not pause_event.is_set():
            if cancel_event.is_set():
                logger("❌ 停止寄送，使用者已取消")
                break
            time.sleep(0.1)
        if cancel_event.is_set():
            break

        try:
            recipient = row["Email"]
            salutation = row["Salutation"]
            statement = random.choice(closing_statements)

            body = html_body.replace("[salutation]", salutation).replace(
                "[statement]", statement
            )

            def repl(match):
                idx = match.group(1)
                if idx == "":
                    return image_html_all
                try:
                    index = int(idx) - 1  # 讓 [image1] 代表第一張圖
                    if index < 0:
                        raise IndexError
                    cid = cid_list[index]
                    return generate_image_html([cid])
                except (ValueError, IndexError):
                    logger(f"⚠️ 無效的圖片佔位符：[image{idx}] → 找不到對應圖片")
                    return ""

            body = re.sub(r"\[image(\d*)\]", repl, body)

            backend.send(
                mode,
                recipient,
                subject,
                body,
                embedded_images,
                real_attachments,
            )
            logger(f"✉ 已處理：{recipient} / {salutation} / {statement}")
            progress_update(i, total, recipient)
            time.sleep(DELAY_SEND if mode == "send" else DELAY_DRAFT)
        except Exception as e:
            logger(f"❌ 寄送失敗：{recipient} - {e}")
            progress_update(i, total, f"{recipient} ❌")

    logger("✅ 所有郵件處理完成")

    if finish_callback:
        finish_callback(last_index, total)
        
    if use_outlook:
        pythoncom.CoUninitialize()

if __name__ == "__main__":
    root = Tk()
    GUI(root)
    root.mainloop()
