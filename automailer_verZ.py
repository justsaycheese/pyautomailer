import extract_msg
import RTFDE.text_extraction as rtf_te
import pandas as pd
import win32com.client as win32
import os
import random
import sys
import time
import logging
from pathlib import Path
from tkinter import (
    Tk, Label, Button, filedialog, StringVar,
    OptionMenu, messagebox, ttk, Text, END, Toplevel, Scrollbar, scrolledtext, Frame
)
import threading
from openpyxl import load_workbook

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# âš™ï¸ Config & Log
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
CLOSING_STATEMENTS = [
    "Thanks & Best Regards", "Kind Regards", "Sincerely",
    "With sincere appreciation", "With gratitude", "Gratefully", "Warm regards"
]
DELAY_SEND = 10
DELAY_DRAFT = 1
LOG_FILE = "automailer_log.txt"
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, format="%(asctime)s - %(message)s")

def patch_rtfde_decode() -> None:
    """Patch RTFDE to ignore undecodable hex characters."""

    def _patched_decode_hex_char(item: bytes, codec: str | None):
        if codec is None:
            codec = "CP1252"
        try:
            return item.decode(codec).encode()
        except UnicodeDecodeError:
            return item.decode(codec, errors="ignore").encode()

    rtf_te.decode_hex_char = _patched_decode_hex_char

patch_rtfde_decode()

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# ğŸ“‚ Utils
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def load_recipients_or_csv(file_path, visible_only=False):
    ext = Path(file_path).suffix.lower()
    if ext == '.csv':
        df = pd.read_csv(file_path)
    elif ext in ['.xls', '.xlsx']:
        if not visible_only:
            df = pd.read_excel(file_path)
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            visible_rows = [
                [cell.value for cell in row]
                for row in ws.iter_rows(min_row=2)
                if not ws.row_dimensions[row[0].row].hidden
            ]
            df = pd.DataFrame(visible_rows, columns=headers)
    else:
        raise ValueError(f"Unsupported file type: {file_path}")

    return df

def validate_recipient_columns(df):
    """Ensure required columns are present in the loaded DataFrame."""
    required = {"Email", "Salutation"}
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Missing column(s): {', '.join(missing)}")

def get_base_dir():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    else:
        return Path(__file__).parent

def load_embeds(embed_dir=None):
    if embed_dir is None:
        embed_dir = get_base_dir() / "embed"
    embed_dir.mkdir(exist_ok=True)
    return {
        f"img_{random.randint(1000,9999)}_{f.stem}": f.resolve()
        for f in embed_dir.glob("*") if f.suffix.lower() in ['.png', '.jpg', '.jpeg', '.gif']
    }

def load_attachments(attachment_dir=None):
    if attachment_dir is None:
        attachment_dir = get_base_dir() / "attachment"
    attachment_dir.mkdir(exist_ok=True)
    return [f.resolve() for f in attachment_dir.glob("*") if f.is_file()]

def generate_image_html(embeds):
    return ''.join(f'<img src="cid:{cid}" style="display:block; margin-bottom:10px;"><br>' for cid in embeds)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# ğŸ–¥ï¸ GUI Class
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
class GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Automailer è‡ªå‹•å¯„ä¿¡å·¥å…·")        

        # æ¨¡å¼åˆ‡æ›ï¼ˆå¯„é€æˆ–å­˜ç¨¿ï¼‰
        self.mode_var = StringVar(value="draft")
        self.recipient_file = ""
        self.exclusion_file = ""
        self.msg_template = ""
        self.embed_dir = None
        self.attachment_dir = None

        self.embed_files = StringVar(value="å°šæœªé¸æ“‡")
        self.attachment_files = StringVar(value="å°šæœªé¸æ“‡")
        self.recipient_label = StringVar(value="å°šæœªé¸æ“‡")
        self.exclusion_label = StringVar(value="å°šæœªé¸æ“‡")
        self.template_label = StringVar(value="å°šæœªé¸æ“‡")
        self.select_mode_var = StringVar(value="è³‡æ–™å¤¾")  # é è¨­ã€Œè³‡æ–™å¤¾ã€æ¨¡å¼
        self.folder_mode = True                          # èˆ‡å‰ä¿æŒä¸€è‡´çš„å¸ƒæ—æ——æ¨™
        self.embed_paths = {}            # å­—å…¸
        self.attachments = []

        # é€²åº¦æ–‡å­—å’Œ Progressbar
        self.progress_label = StringVar(value="")
        self.progress_bar = None

        # æ—¥èªŒè¦–çª—ç›¸é—œ
        self.log_window = None
        self.log_buffer = ["âœ… ç¨‹å¼å·²å•Ÿå‹•"]

        # â”€â”€â”€ pause_event & cancel_event â”€â”€â”€
        self.pause_event = threading.Event()
        self.pause_event.set()  # ä¸€é–‹å§‹ç‚ºã€Œå·²è¨­å®šã€ï¼Œä»£è¡¨ä¸æš«åœ

        self.cancel_event = threading.Event()  # ä¸€é–‹å§‹ç‚º Falseï¼Œä»£è¡¨æœªå–æ¶ˆ
        
        
        # â€”â€”å–å¾— Outlook Accounts â€”â€” 
        outlook_app = win32.Dispatch('Outlook.Application')
        session = outlook_app.GetNamespace("MAPI")
        accounts = [acct.DisplayName for acct in session.Accounts]

        # å¦‚æœåªæœ‰ä¸€å€‹å¸³æˆ¶ï¼Œä¹ŸæŠŠå®ƒæ”¾é€²å»
        if not accounts:
            accounts = ["(No Account Found)"]
        self.account_var = StringVar(root)
        self.account_var.set(accounts[0])

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ UI Frames â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        mode_frame = Frame(root, pady=5, padx=5, relief="groove", borderwidth=2)
        mode_frame.grid(row=0, column=0, columnspan=2, sticky="EW")
        Label(mode_frame, text="å¯„ä»¶å¸³æˆ¶ï¼š").grid(row=0, column=0, sticky="w", pady=5)
        
        self.account_menu = OptionMenu(mode_frame, self.account_var, *accounts)
        self.account_menu.grid(row=0, column=1, sticky="W", pady=5)
        self.account_menu.config(width=20)
        
        Label(mode_frame, text="é¸æ“‡å¯„é€æ¨¡å¼:").grid(row=1, column=0, sticky="W")
        mode_menu = OptionMenu(mode_frame, self.mode_var, "send", "draft")
        mode_menu.config(width=5)
        mode_menu.grid(row=1, column=1, sticky="W")

        file_frame = Frame(root, pady=5, padx=5, relief="groove", borderwidth=2)
        file_frame.grid(row=1, column=0, columnspan=2, sticky="EW")
        self.embed_btn = Button(file_frame, text="ğŸ–¼ é¸æ“‡åœ–ç‰‡è³‡æ–™å¤¾", command=self.select_embed, width=20)
        self.embed_btn.grid(row=1, column=0, pady=5)
        Label(file_frame, textvariable=self.embed_files, wraplength=270, justify="left").grid(row=1, column=1, sticky="W")

        self.attachment_btn = Button(file_frame, text="ğŸ“ é¸æ“‡é™„ä»¶è³‡æ–™å¤¾", command=self.select_attachment, width=20)
        self.attachment_btn.grid(row=2, column=0, pady=5)
        Label(file_frame, textvariable=self.attachment_files, wraplength=270, justify="left").grid(row=2, column=1, sticky="W")

        inner_frame1 = Frame(file_frame, borderwidth=0)
        inner_frame1.grid(row=0, column=0, columnspan=2, sticky="EW")
        Label(inner_frame1, text="é¸å–æ¨¡å¼ï¼š").grid(row=0, column=0, sticky="W", pady=5)
        mode_select = OptionMenu(inner_frame1, self.select_mode_var,"è³‡æ–™å¤¾", "å¤šæª”æ¡ˆ", command=self.on_select_mode)
        mode_select.config(width=8)
        mode_select.grid(row=0, column=1, sticky="W", pady=5)

        choose_frame = Frame(root, pady=5, padx=5, relief="groove", borderwidth=2)
        choose_frame.grid(row=2, column=0, columnspan=2, sticky="EW")
        Button(choose_frame, text="ğŸ“‹ é¸æ“‡æ”¶ä»¶äºº", command=self.load_recipients, width=20).grid(row=0, column=0, pady=5)
        Label(choose_frame, textvariable=self.recipient_label, wraplength=270, justify="left").grid(row=0, column=1, sticky="W")
        Button(choose_frame, text="ğŸš« é¸æ“‡æ’é™¤æ¸…å–®", command=self.load_exclusions, width=20).grid(row=1, column=0, pady=5)
        Label(choose_frame, textvariable=self.exclusion_label, wraplength=270, justify="left").grid(row=1, column=1, sticky="W")
        Button(choose_frame, text="âœ‰ é¸æ“‡éƒµä»¶ç¯„æœ¬", command=self.load_msg_template, width=20).grid(row=2, column=0, pady=5)
        Label(choose_frame, textvariable=self.template_label, wraplength=270, justify="left").grid(row=2, column=1, sticky="W")

        Button(root, text="ğŸš€ é–‹å§‹å¯„ä¿¡", command=self.start_process).grid(row=3, column=0, pady=10)
        Button(root, text="ğŸªµ æŸ¥çœ‹æ—¥èªŒ", command=self.show_log_window).grid(row=3, column=1)

        Label(root, text="çµå°¾è© (ä¸€è¡Œä¸€å€‹)").grid(row=4, column=0, columnspan=2)
        self.closing_text = scrolledtext.ScrolledText(root, height=7, width=50)
        self.closing_text.grid(row=5, column=0, columnspan=2)
        self.closing_text.insert(END, "\n".join(CLOSING_STATEMENTS))

        # â”€â”€â”€ Pause/Resume æŒ‰éˆ• & Cancel æŒ‰éˆ•ï¼ˆä¸€é–‹å§‹å…ˆæ”¾ä½ç½®ï¼Œå†éš±è—ï¼‰ â”€â”€â”€
        self.pause_button = Button(root, text="æš«åœ", font=('Arial',12,'bold'), fg='#f00' , state="normal", command=self.toggle_pause, width=10, height=1)
        self.pause_button.grid(row=7, column=0, pady=5)
        self.pause_button.grid_remove()   # å…ˆéš±è—

        self.cancel_button = Button(root, text="å–æ¶ˆ", font=('Arial',12,'bold'), fg='#f00', state="normal", command=self.cancel_process, width=10, height=1)
        self.cancel_button.grid(row=7, column=1, pady=5)
        self.cancel_button.grid_remove()  # å…ˆéš±è—

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ é€²åº¦å€å¡Š â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        Label(root, textvariable=self.progress_label).grid(row=9, column=0, columnspan=2, pady=5, sticky="S")
        self.progress_bar = ttk.Progressbar(root, length=300, mode='determinate')
        self.progress_bar.grid(row=10, column=0, columnspan=2, pady=5)
    
    def on_select_mode(self, choice):
        """ç•¶ OptionMenu è®Šå‹•æ™‚å‘¼å«ï¼›åŒæ­¥æ›´æ–° folder_mode èˆ‡æŒ‰éˆ•æ–‡å­—"""
        self.folder_mode = (choice == "è³‡æ–™å¤¾")  # True=è³‡æ–™å¤¾æ¨¡å¼

        if self.folder_mode:
            self.embed_btn.config(text="ğŸ–¼ é¸æ“‡åœ–ç‰‡è³‡æ–™å¤¾")
            self.attachment_btn.config(text="ğŸ“ é¸æ“‡é™„ä»¶è³‡æ–™å¤¾")
        else:
            self.embed_btn.config(text="ğŸ–¼ é¸æ“‡åœ–ç‰‡æª”æ¡ˆ")
            self.attachment_btn.config(text="ğŸ“ é¸æ“‡é™„ä»¶æª”æ¡ˆ")


    def select_embed(self):
        if self.folder_mode:  # â–¸ è³‡æ–™å¤¾æ¨¡å¼
            path = filedialog.askdirectory(title="é¸æ“‡åµŒå…¥åœ–ç‰‡è³‡æ–™å¤¾")
            if not path:
                return
            self.embed_dir = Path(path)
            self.embed_paths = load_embeds(self.embed_dir)
        else:  # â–¸ å¤šæª”æ¡ˆæ¨¡å¼
            paths = filedialog.askopenfilenames(
                title="é¸æ“‡åœ–ç‰‡æª”æ¡ˆ",
                filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif")]
            )
            if not paths:
                return
            # ç”¢ç”Ÿè·Ÿ load_embeds ç›¸åŒçµæ§‹çš„ dict
            self.embed_paths = {
                f"img_{random.randint(1000, 9999)}_{Path(p).stem}": Path(p)
                for p in paths
            }

        file_names = [p.name for p in self.embed_paths.values()]
        self.embed_files.set(", ".join(file_names) or "ç„¡æª”æ¡ˆ")
        self.log(f"âœ… å·²è¼‰å…¥ {len(self.embed_paths)} å¼µåœ–ç‰‡")

    def select_attachment(self):
        if self.folder_mode:
            # â–¸ è³‡æ–™å¤¾æ¨¡å¼
            path = filedialog.askdirectory(title="é¸æ“‡é™„ä»¶è³‡æ–™å¤¾")
            if not path:
                return
            self.attachment_dir = Path(path)
            self.attachments = load_attachments(self.attachment_dir)
        else:
            # â–¸ å¤šæª”æ¡ˆæ¨¡å¼
            paths = filedialog.askopenfilenames(title="é¸æ“‡é™„ä»¶æª”æ¡ˆ")
            if not paths:
                return
            self.attachments = [Path(p) for p in paths]

        file_names = [p.name for p in self.attachments]
        self.attachment_files.set(", ".join(file_names) or "ç„¡æª”æ¡ˆ")
        self.log(f"âœ… å·²è¼‰å…¥ {len(self.attachments)} å€‹é™„ä»¶")

    def load_recipients(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv")])
        if path:
            self.recipient_file = path
            self.recipient_label.set(Path(path).name)

    def load_exclusions(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv")])
        if path:
            self.exclusion_file = path
            self.exclusion_label.set(Path(path).name)

    def load_msg_template(self):
        path = filedialog.askopenfilename(filetypes=[("MSG Files", "*.msg")])
        if path:
            self.msg_template = path
            self.template_label.set(Path(path).name)

    def show_log_window(self):
        if self.log_window and self.log_window.winfo_exists():
            self.log_window.lift()
            return
        self.log_window = Toplevel(self.root)
        self.log_window.title("ğŸªµ æ—¥èªŒç´€éŒ„")

        frame = Frame(self.log_window)
        frame.pack(fill="both", expand=True)

        self.log_text = Text(frame, wrap="word", font=("Courier", 9))
        scrollbar = Scrollbar(frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        for msg in self.log_buffer:
            self.log_text.insert(END, msg + "\n")
        self.log_text.config(state="disabled")

        clear_btn = Button(frame, text="ğŸ§¹æ¸…ç©º", command=self.clear_log, relief="groove", bg="lightgray")
        clear_btn.place(relx=1.0, rely=0.0, anchor="ne", x=1, y=0)

    def clear_log(self):
        self.log_buffer.clear()
        if self.log_window and hasattr(self, "log_text") and self.log_window.winfo_exists():
            self.log_text.config(state="normal")
            self.log_text.delete("1.0", END)
            self.log_text.config(state="disabled")
        with open(LOG_FILE, "w") as f:
            f.write("")
        self.log("ğŸ§¹ æ—¥èªŒå·²æ¸…ç©º")

    def log(self, msg):
        logging.info(msg)
        self.log_buffer.append(msg)
        if self.log_window and hasattr(self, "log_text") and self.log_window.winfo_exists():
            def append_log():
                self.log_text.config(state="normal")
                self.log_text.insert(END, msg + "\n")
                self.log_text.see(END)
                self.log_text.config(state="disabled")
            self.root.after(0, append_log)

    def start_process(self):
        # â”€â”€â”€ é‡æ–°é–‹å§‹æ™‚ï¼Œè¦å…ˆé‡ç½®é€²åº¦æ¨™ç±¤èˆ‡é€²åº¦æ¢ â”€â”€â”€
        self.progress_label.set("")
        self.progress_bar['value'] = 0

        global CLOSING_STATEMENTS
        user_input = self.closing_text.get("1.0", END).strip().splitlines()
        CLOSING_STATEMENTS = [line.strip() for line in user_input if line.strip()]
        if not self.recipient_file or not self.msg_template:
            messagebox.showerror("éŒ¯èª¤", "è«‹é¸æ“‡æ”¶ä»¶äººæ¸…å–®å’Œéƒµä»¶ç¯„æœ¬")
            return
            
        # -------- ä¾æ¨¡å¼æ•´ç†åœ–ç‰‡èˆ‡é™„ä»¶æ¸…å–® --------
        if self.embed_paths:                    # â†– å¤šæª”æ¡ˆæ¨¡å¼
            embedded_images = self.embed_paths
        elif self.embed_dir is not None:        # â†– è³‡æ–™å¤¾æ¨¡å¼
            embedded_images = load_embeds(self.embed_dir)
        else:
            embedded_images = load_embeds()

        if self.attachments:                    # â†– å¤šæª”æ¡ˆæ¨¡å¼
            real_attachments = self.attachments
        elif self.attachment_dir is not None:   # â†– è³‡æ–™å¤¾æ¨¡å¼
            real_attachments = load_attachments(self.attachment_dir)
        else:
            real_attachments = load_attachments()

        embed_list = "\n".join([f"- {cid} â†’ {p.name}" for cid, p in embedded_images.items()]) or "ç„¡"
        attachment_list = "\n".join([f"- {p.name}" for p in real_attachments]) or "ç„¡"
        statement_list = "\n".join(CLOSING_STATEMENTS)
        confirm_message = f"""ğŸ“‚ æª¢æŸ¥å®Œç•¢ï¼Œæº–å‚™å¯„ä¿¡ï¼š{self.mode_var.get()}

å¯„ä»¶å¸³æˆ¶ï¼š{self.account_var.get()}

åµŒå…¥åœ–ç‰‡:
{embed_list}

é™„åŠ æª”æ¡ˆ:
{attachment_list}

çµå°¾è©:
{statement_list}

æ˜¯å¦ç¢ºèªé–‹å§‹å¯„ä¿¡ï¼Ÿ
"""
        if not messagebox.askyesno("ç¢ºèªå¯„ä¿¡", confirm_message):
            self.progress_label.set("â›” æ“ä½œå–æ¶ˆ")
            self.log("â›” æ“ä½œå–æ¶ˆ")
            return

        # åœ¨é–‹å§‹ä¹‹å‰ï¼Œé‡ç½® cancel_event ä¸¦è¨­å®š pause_event
        self.cancel_event.clear()
        self.pause_event.set()

        # é¡¯ç¤ºæš«åœå’Œå–æ¶ˆæŒ‰éˆ•
        self.pause_button.grid()   # å¾éš±è—ç‹€æ…‹æ¢å¾©
        self.pause_button.config(text="æš«åœ")
        self.cancel_button.grid()  # å¾éš±è—ç‹€æ…‹æ¢å¾©

        # å•Ÿå‹•èƒŒæ™¯åŸ·è¡Œç·’ï¼Œå‚³å…¥ pause_event å’Œ cancel_event
        threading.Thread(
            target=run_automailer,
            args=(
                self.mode_var.get(),
                self.recipient_file,
                self.exclusion_file,
                self.msg_template,
                self.update_progress,
                self.log,
                embedded_images,          # â† æ”¹å‚³ã€Œæœ€çµ‚ dictã€
                real_attachments,         # â† æ”¹å‚³ã€Œæœ€çµ‚ listã€
                self.pause_event,
                self.cancel_event,
                self.on_finish,
                self.account_var.get()
            ),
            daemon=True
        ).start()

    def toggle_pause(self):
        """åˆ‡æ›æš«åœ / ç¹¼çºŒ ç‹€æ…‹ï¼Œä¸¦æ›´æ–°æŒ‰éˆ•æ–‡å­—ã€‚"""
        if self.pause_event.is_set():
            # ç”±ã€Œå¯åŸ·è¡Œã€è®Šæˆã€Œæš«åœã€
            self.pause_event.clear()
            self.pause_button.config(text="ç¹¼çºŒ")
            self.log("â¸ï¸ å·²æš«åœå¯„é€")
        else:
            # ç”±ã€Œæš«åœã€è®Šæˆã€Œå¯åŸ·è¡Œã€
            self.pause_event.set()
            self.pause_button.config(text="æš«åœ")
            self.log("â–¶ï¸ å·²ç¹¼çºŒå¯„é€")

    def cancel_process(self):
        """ä½¿ç”¨è€…é»ã€Œå–æ¶ˆã€æ™‚ï¼Œè§¸ç™¼ cancel_eventï¼Œä¸¦éš±è—æŒ‰éˆ•ã€‚"""
        # è¨­å®š cancel_eventï¼Œè®“ run_automailer è¿´åœˆè·³å‡º
        self.cancel_event.set()
        self.log("âŒ ä½¿ç”¨è€…å·²å–æ¶ˆå¯„é€")
        # éš±è—æŒ‰éˆ•
        self.pause_button.grid_remove()
        self.cancel_button.grid_remove()
        # æ›´æ–°é€²åº¦æ–‡å­—ç‚ºå·²å–æ¶ˆ
        self.progress_label.set("âŒ å·²å–æ¶ˆå¯„é€")

    def update_progress(self, index, total, current_email):
        pct = int((index + 1) / total * 100)
        self.progress_label.set(f"{pct}% - è™•ç† {index + 1}/{total}: {current_email}")
        self.progress_bar['value'] = pct
        self.root.update_idletasks()
        
    def on_finish(self, last_index, total):
        """æµç¨‹è·‘å®Œå¾Œï¼ŒæŠŠæš«åœèˆ‡å–æ¶ˆæŒ‰éˆ•éš±è—æ‰ã€‚"""
        self.pause_button.grid_remove()
        self.cancel_button.grid_remove()
        # å¯ä»¥æ›´æ–°é€²åº¦æ–‡å­—è¡¨é”ã€Œå·²å®Œæˆã€ï¼š
        finished_count = last_index + 1 if last_index is not None else total
        self.progress_label.set(f"âœ… å…¨éƒ¨å¯„é€å®Œæˆ {finished_count}/{total}")

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# ğŸš€ Email Sending Logic
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def run_automailer(mode, recipients_path, exclusion_path, msg_template_path,
                   progress_update, logger, embedded_images, real_attachments, pause_event, cancel_event, finish_callback, send_account_name):
    
    outlook = win32.Dispatch('Outlook.Application')
    session = outlook.GetNamespace("MAPI")
    # æ‰¾å‡ºä½¿ç”¨è€…é¸çš„é‚£å€‹ Account ç‰©ä»¶
    send_account = None
    for acct in session.Accounts:
        if acct.DisplayName == send_account_name:
            send_account = acct
            break
    if send_account is None:
        logger(f"âš ï¸ æ‰¾ä¸åˆ°å¸³æˆ¶ {send_account_name}ï¼Œå°‡ä½¿ç”¨é è¨­å¸³æˆ¶ã€‚")
    
    try:
        recipients = load_recipients_or_csv(recipients_path, visible_only=True)
        validate_recipient_columns(recipients)
    except ValueError as e:
        messagebox.showerror("æª”æ¡ˆéŒ¯èª¤", str(e))
        logger(f"æ”¶ä»¶äººæ¸…å–®éŒ¯èª¤: {e}")
        if finish_callback:
            finish_callback(None, 0)
        return
    exclusion_emails = []
    if exclusion_path and os.path.exists(exclusion_path):
        try:
            exclusion_df = load_recipients_or_csv(exclusion_path)
            exclusion_emails = exclusion_df['Email'].tolist()
        except Exception as e:
            logger(f"æ’é™¤æ¸…å–®è®€å–å¤±æ•—: {e}")
    filtered = recipients[~recipients['Email'].isin(exclusion_emails)]

    msg = extract_msg.Message(msg_template_path)
    subject = msg.subject
    try:
        raw_html_body = msg.htmlBody
    except UnicodeDecodeError as e:
        logger(f"HTML è§£æå¤±æ•—: {e}")
        raw_html_body = None
    html_body = raw_html_body.decode('utf-8', errors='ignore') if isinstance(raw_html_body, bytes) else (raw_html_body or '')

    image_html = generate_image_html(embedded_images)

    total = len(filtered)
    
    
    """
    æ–°å¢åƒæ•¸ cancel_eventã€‚æ¯æ¬¡è¿´åœˆé–‹å§‹å‰æˆ– pause æ™‚ï¼Œéƒ½è¦æª¢æŸ¥ cancel_event 
    æ˜¯å¦å·²è¢«è¨­ç½®ã€‚è¨­ç½®å°±ç›´æ¥çµæŸæ•´å€‹æµç¨‹ã€‚
    æ–°å¢ä¸€å€‹åƒæ•¸ pause_eventï¼šthreading.Event ç‰©ä»¶ã€‚
    åœ¨æ¯æ¬¡å¯¦éš›è¦ç™¼é€/å­˜ç¨¿ä¹‹å‰ï¼Œéƒ½å…ˆå‘¼å« pause_event.wait()ã€‚
    ç•¶ pause_event è¢« clear æ™‚ï¼Œwait() æœƒé˜»å¡ï¼›è¢« set æ™‚ç¹¼çºŒåŸ·è¡Œã€‚
    """
    last_index = None
    for i, row in filtered.iterrows():
        last_index = i
        # è‹¥ä½¿ç”¨è€…æŒ‰äº†ã€Œå–æ¶ˆã€ï¼Œå°±ç›´æ¥è·³å‡º
        if cancel_event.is_set():
            logger("âŒ åœæ­¢å¯„é€ï¼Œä½¿ç”¨è€…å·²å–æ¶ˆ")
            break

        # æš«åœè™•ç†ï¼šè‹¥ pause_event æ²’è¢« setï¼Œå°±æŒçºŒå°ç¡ä¸¦æ¯æ¬¡æª¢æŸ¥ cancel_event
        while not pause_event.is_set():
            if cancel_event.is_set():
                logger("âŒ åœæ­¢å¯„é€ï¼Œä½¿ç”¨è€…å·²å–æ¶ˆ")
                break
            time.sleep(0.1)
        if cancel_event.is_set():
            break

        try:
            recipient = row["Email"]
            salutation = row["Salutation"]
            statement = random.choice(CLOSING_STATEMENTS)

            body = html_body.replace("[salutation]", salutation)\
                            .replace("[statement]", statement)\
                            .replace("[image]", image_html)

            mail = outlook.CreateItem(0)
            if send_account:
                # ç›´æ¥æŒ‡å®šè¦ç”¨å“ªå€‹å¸³æˆ¶
                mail._oleobj_.Invoke(64209, 0, 8, 1, send_account)
            mail.Subject = subject
            mail.To = recipient
            mail.HTMLBody = body

            for cid, path in embedded_images.items():
                a = mail.Attachments.Add(Source=str(path))
                a.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)

            for file_path in real_attachments:
                mail.Attachments.Add(Source=str(file_path))

            if mode == "send":
                mail.Send()
                logger(f"âœ‰ å·²ç™¼é€ï¼š{recipient} / {salutation} / {statement}")
                progress_update(i, total, recipient)
                time.sleep(DELAY_SEND)
            else:
                mail.Save()
                logger(f"ğŸ“ è‰ç¨¿å„²å­˜ï¼š{recipient} / {salutation} / {statement}")
                progress_update(i, total, recipient)
                time.sleep(DELAY_DRAFT)

            del mail
        except Exception as e:
            logger(f"âŒ å¯„é€å¤±æ•—ï¼š{recipient} - {e}")
            progress_update(i, total, f"{recipient} âŒ")

    del outlook
    logger("âœ… æ‰€æœ‰éƒµä»¶è™•ç†å®Œæˆ")
    
    if finish_callback:
        finish_callback(last_index, total)

if __name__ == "__main__":
    root = Tk()
    GUI(root)
    root.mainloop()
