"""Utility helpers for automailer."""
import json
import logging
import re
import uuid
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import RTFDE.text_extraction as rtf_te


def patch_rtfde_decode() -> None:
    """Patch RTFDE to ignore undecodable hex characters."""

    def _patched_decode_hex_char(item: bytes, codec: str | None):
        if codec is None:
            codec = "CP1252"
        try:
            return item.decode(codec).encode()
        except UnicodeDecodeError:
            return item.decode(codec, errors="ignore").encode()

    rtf_te.decode_hex_char = _patched_decode_hex_char


patch_rtfde_decode()


def load_recipients_or_csv(file_path, visible_only=False):
    ext = Path(file_path).suffix.lower()
    if ext == ".csv":
        df = pd.read_csv(file_path)
    elif ext in [".xls", ".xlsx"]:
        if not visible_only:
            df = pd.read_excel(file_path)
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            visible_rows = [
                [cell.value for cell in row]
                for row in ws.iter_rows(min_row=2)
                if not ws.row_dimensions[row[0].row].hidden
            ]
            df = pd.DataFrame(visible_rows, columns=headers)
    else:
        raise ValueError(f"Unsupported file type: {file_path}")

    return df


def validate_recipient_columns(df):
    """Ensure required columns are present in the loaded DataFrame."""
    required = {"Email", "Salutation"}
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Missing column(s): {', '.join(missing)}")


def get_base_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    else:
        return Path(__file__).parent


SETTINGS_FILE = get_base_dir() / "settings.json"


def load_settings_file():
    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Failed to load settings: {e}")
    return {}


def save_settings_file(data):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logging.error(f"Failed to save settings: {e}")


def safe_cid(stem: str) -> str:
    """Return a filesystem safe CID for embeds."""
    clean = re.sub(r"[^A-Za-z0-9_-]+", "_", stem)
    return f"{uuid.uuid4().hex[:8]}_{clean}"


def load_embeds(embed_dir):
    if embed_dir is None:
        return {}
    embed_dir = Path(embed_dir)
    return {
        safe_cid(f.stem): f.resolve()
        for f in embed_dir.glob("*")
        if f.suffix.lower() in [".png", ".jpg", ".jpeg", ".gif"]
    }


def load_attachments(attachment_dir):
    if attachment_dir is None:
        return []
    attachment_dir = Path(attachment_dir)
    return [f.resolve() for f in attachment_dir.glob("*") if f.is_file()]


def generate_image_html(embeds):
    return "".join(
        f'<img src="cid:{cid}" style="display:block; margin-bottom:10px;"/><br>'
        for cid in embeds
    )
